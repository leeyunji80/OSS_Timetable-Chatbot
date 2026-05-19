from __future__ import annotations

import argparse
import csv
import logging
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


COLUMNS = [
    "개설 연도",
    "개설 학과",
    "수강 대상",
    "교과목 번호",
    "분반 번호",
    "교과목명",
    "이수구분",
    "학점",
    "이론",
    "실습",
    "수업방식",
    "요일",
    "교시",
    "강의실",
    "담당교수",
    "강의 정원",
    "방법_강의(%)",
    "방법_토의토론(%)",
    "방법_실험실습(%)",
    "방법_현장학습(%)",
    "방법_발표(%)",
    "방법_기타(%)",
    "평가_중간(%)",
    "평가_기말(%)",
    "평가_출석(%)",
    "평가_퀴즈(%)",
    "평가_과제(%)",
    "평가_기타(%)",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t\r\f\v]+", " ", str(value)).strip()


def as_cell(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


@dataclass(frozen=True)
class CellHit:
    row: int
    col: int
    value: Any

    @property
    def coord(self) -> str:
        return as_cell(self.row, self.col)


class LectureExcelReader:
    """Read Chungbuk National University lecture XLSX files by cell coordinates.

    The university files sometimes contain ``applyNumberForm`` in styles.xml.
    openpyxl expects ``applyNumberFormat``. This reader first tries to open the
    workbook normally, then patches only that typo in memory and retries.
    """

    def __init__(self, path: str | Path, sheet_name: str | None = None, logger: logging.Logger | None = None):
        self.path = Path(path)
        self.logger = logger or logging.getLogger(self.__class__.__name__)
        self.workbook = self._load_workbook()
        self.sheet = self.workbook[sheet_name] if sheet_name else self.workbook[self.workbook.sheetnames[0]]
        self._merged_ranges = list(self.sheet.merged_cells.ranges)

    def _load_workbook(self):
        try:
            return load_workbook(self.path, data_only=True)
        except TypeError as exc:
            if "applyNumberForm" not in str(exc):
                raise
            self.logger.debug("styles.xml typo found; patching applyNumberForm in memory: %s", self.path)
            return load_workbook(self._fixed_xlsx_bytes(), data_only=True)

    def _fixed_xlsx_bytes(self) -> BytesIO:
        output = BytesIO()
        with ZipFile(self.path, "r") as zin, ZipFile(output, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = data.replace(b"applyNumberForm=", b"applyNumberFormat=")
                zout.writestr(item, data)
        output.seek(0)
        return output

    def raw_value(self, row: int, col: int) -> Any:
        return self.sheet.cell(row=row, column=col).value

    def merged_range_for(self, row: int, col: int):
        for merged in self._merged_ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return merged
        return None

    def merged_top_left(self, row: int, col: int) -> tuple[int, int]:
        merged = self.merged_range_for(row, col)
        if merged:
            return merged.min_row, merged.min_col
        return row, col

    def cell_value(self, row: int, col: int, fill_merged: bool = True) -> Any:
        if not fill_merged:
            return self.raw_value(row, col)
        top_row, top_col = self.merged_top_left(row, col)
        return self.raw_value(top_row, top_col)

    def find_label(self, label: str, exact: bool = False) -> list[CellHit]:
        needle = normalize_text(label)
        hits: list[CellHit] = []
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                value = self.raw_value(row, col)
                if value is None:
                    continue
                haystack = normalize_text(value)
                matched = haystack == needle if exact else needle in haystack
                if matched:
                    hits.append(CellHit(row, col, value))
        return hits

    def first_label(self, label: str, exact: bool = False) -> CellHit | None:
        hits = self.find_label(label, exact=exact)
        return hits[0] if hits else None

    def read_values_to_right(
        self,
        label: str,
        *,
        exact: bool = False,
        stop_labels: Iterable[str] = (),
        max_scan_cols: int = 25,
    ) -> list[CellHit]:
        hit = self.first_label(label, exact=exact)
        if not hit:
            self.logger.warning("label not found: %s", label)
            return []

        label_range = self.merged_range_for(hit.row, hit.col)
        start_col = (label_range.max_col + 1) if label_range else (hit.col + 1)
        stop_needles = [normalize_text(item) for item in stop_labels]
        values: list[CellHit] = []

        for col in range(start_col, min(self.sheet.max_column, start_col + max_scan_cols - 1) + 1):
            value = self.raw_value(hit.row, col)
            if value is None:
                continue
            normalized = normalize_text(value)
            if any(stop and stop in normalized for stop in stop_needles):
                break
            values.append(CellHit(hit.row, col, value))
        self.logger.debug(
            "read_values_to_right(%r): label %s -> %s",
            label,
            hit.coord,
            ", ".join(f"{v.coord}={v.value!r}" for v in values),
        )
        return values

    def read_first_to_right(self, label: str, *, exact: bool = False, stop_labels: Iterable[str] = ()) -> CellHit | None:
        values = self.read_values_to_right(label, exact=exact, stop_labels=stop_labels)
        return values[0] if values else None

    def row_cells(self, row: int, *, fill_merged: bool = False) -> list[Any]:
        return [self.cell_value(row, col, fill_merged=fill_merged) for col in range(1, self.sheet.max_column + 1)]

    def all_rows(self, *, fill_merged: bool = False) -> list[list[Any]]:
        return [self.row_cells(row, fill_merged=fill_merged) for row in range(1, self.sheet.max_row + 1)]

    def dump_all_rows(self, *, limit: int | None = None, fill_merged: bool = False) -> str:
        lines: list[str] = []
        max_row = min(self.sheet.max_row, limit or self.sheet.max_row)
        for row in range(1, max_row + 1):
            parts = []
            for col in range(1, self.sheet.max_column + 1):
                value = self.cell_value(row, col, fill_merged=fill_merged)
                if value is not None:
                    parts.append(f"{as_cell(row, col)}={value!r}")
            if parts:
                lines.append(f"row {row}: " + " | ".join(parts))
        return "\n".join(lines)

    def dump_structure(self) -> str:
        merged = ", ".join(str(rng) for rng in self._merged_ranges)
        return (
            f"file: {self.path}\n"
            f"sheet: {self.sheet.title}\n"
            f"dimension: {self.sheet.max_row} x {self.sheet.max_column}\n"
            f"merged_ranges: {merged}"
        )


class LectureParser:
    METHOD_HEADERS = {
        "강의": "방법_강의(%)",
        "토의/토론": "방법_토의토론(%)",
        "실험/실습": "방법_실험실습(%)",
        "현장학습": "방법_현장학습(%)",
        "개별/팀별발표": "방법_발표(%)",
        "기타": "방법_기타(%)",
    }

    EVALUATION_HEADERS = {
        "중간고사": "평가_중간(%)",
        "기말고사": "평가_기말(%)",
        "출석": "평가_출석(%)",
        "퀴즈": "평가_퀴즈(%)",
        "과제": "평가_과제(%)",
        "기타": "평가_기타(%)",
    }

    def __init__(self, logger: logging.Logger | None = None):
        self.logger = logger or logging.getLogger(self.__class__.__name__)

    def parse_file(self, path: str | Path) -> dict[str, Any]:
        reader = LectureExcelReader(path, logger=self.logger)
        row = {column: "" for column in COLUMNS}

        row["개설 연도"] = self._parse_year(self._read_first(reader, "개설연도-학기", stop_labels=["개설학과"]))
        row["개설 학과"] = self._read_first(reader, "개설학과")
        row["수강 대상"] = self._read_first(reader, "수강대상")

        course_parts = reader.read_values_to_right("교과목번호-분반번호", stop_labels=["교과목명"])
        row["교과목 번호"] = clean_text(course_parts[0].value) if len(course_parts) >= 1 else ""
        row["분반 번호"] = clean_text(course_parts[1].value) if len(course_parts) >= 2 else ""
        self._log_field("교과목 번호", course_parts[0] if len(course_parts) >= 1 else None)
        self._log_field("분반 번호", course_parts[1] if len(course_parts) >= 2 else None)

        row["교과목명"] = self._read_first(reader, "교과목명")
        row["이수구분"] = self._read_first(reader, "이수구분")

        credit_text = self._read_first(reader, "학점/시수")
        row["학점"], row["이론"], row["실습"] = self._parse_credit_hours(credit_text)

        row["수업방식"] = self._read_first(reader, "수업방식")
        days, periods, rooms = self._parse_time_room(self._read_first(reader, "강의시간/강의실"))
        row["요일"] = days
        row["교시"] = periods
        row["강의실"] = rooms

        row["담당교수"] = self._read_first(reader, "담당교수")
        row["강의 정원"] = self._read_first(reader, "강의정원")

        row.update(self._parse_percent_table(reader, self.METHOD_HEADERS, section_name="수업진행방법"))
        row.update(self._parse_percent_table(reader, self.EVALUATION_HEADERS, section_name="평가방법"))
        return row

    def _read_first(self, reader: LectureExcelReader, label: str, *, stop_labels: Iterable[str] = ()) -> str:
        hit = reader.read_first_to_right(label, stop_labels=stop_labels)
        self._log_field(label, hit)
        return clean_text(hit.value) if hit else ""

    def _log_field(self, field: str, hit: CellHit | None) -> None:
        if hit:
            self.logger.debug("%s <- %s = %r", field, hit.coord, hit.value)
        else:
            self.logger.debug("%s <- NOT FOUND", field)

    @staticmethod
    def _parse_year(text: str) -> str:
        match = re.search(r"(20\d{2}|\d{4})", text or "")
        return match.group(1) if match else clean_text(text)

    @staticmethod
    def _parse_credit_hours(text: str) -> tuple[str, str, str]:
        numbers = re.findall(r"\d+(?:\.\d+)?", text or "")
        if len(numbers) >= 3:
            return numbers[0], numbers[1], numbers[2]
        if len(numbers) == 1:
            return numbers[0], "", ""
        return "", "", ""

    @staticmethod
    def _parse_percent(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            number = float(value)
            if 0 < number <= 1:
                number *= 100
            return str(int(number)) if number.is_integer() else str(number)

        text = clean_text(value)
        if not text or text == "%" or text == "상세정보":
            return ""
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return ""
        number = float(match.group(0))
        if "%" in text and 0 < number <= 1:
            number *= 100
        return str(int(number)) if number.is_integer() else str(number)

    def _parse_percent_table(
        self,
        reader: LectureExcelReader,
        header_map: dict[str, str],
        *,
        section_name: str,
    ) -> dict[str, str]:
        result = {column: "0" for column in header_map.values()}
        header_row = self._find_percent_header_row(reader, header_map, section_name)
        if not header_row:
            self.logger.warning("percent header row not found: %s", section_name)
            return result

        for col in range(1, reader.sheet.max_column + 1):
            value = reader.raw_value(header_row, col)
            if value is None:
                continue
            header_key = normalize_text(value)
            if header_key not in header_map:
                continue
            output_column = header_map[header_key]
            percent, coord = self._read_percent_below_header(reader, header_row, col)
            result[output_column] = percent if percent != "" else "0"
            self.logger.debug(
                "%s / %s: header %s=%r -> value %s=%r",
                section_name,
                output_column,
                as_cell(header_row, col),
                value,
                coord or "NOT FOUND",
                result[output_column],
            )
        return result

    def _find_percent_header_row(
        self,
        reader: LectureExcelReader,
        header_map: dict[str, str],
        section_name: str,
    ) -> int | None:
        section_hit = reader.first_label(section_name)
        if section_hit:
            search_start = max(1, section_hit.row - 3)
            search_end = min(reader.sheet.max_row, section_hit.row + 4)
        else:
            search_start = 1
            search_end = reader.sheet.max_row

        candidates: list[tuple[int, int]] = []
        header_keys = set(header_map.keys())
        for row in range(search_start, search_end + 1):
            found = 0
            for col in range(1, reader.sheet.max_column + 1):
                value = reader.raw_value(row, col)
                if value is not None and normalize_text(value) in header_keys:
                    found += 1
            if found:
                candidates.append((found, row))

        if not candidates and section_name == "평가방법":
            for row in range(1, reader.sheet.max_row + 1):
                found = 0
                for col in range(1, reader.sheet.max_column + 1):
                    value = reader.raw_value(row, col)
                    if value is not None and normalize_text(value) in header_keys:
                        found += 1
                if found:
                    candidates.append((found, row))

        if not candidates:
            return None
        candidates.sort(reverse=True)
        chosen = candidates[0][1]
        self.logger.debug("%s header row -> %s (matched %s headers)", section_name, chosen, candidates[0][0])
        return chosen

    def _read_percent_below_header(self, reader: LectureExcelReader, header_row: int, header_col: int) -> tuple[str, str]:
        header_range = reader.merged_range_for(header_row, header_col)
        min_col = header_range.min_col if header_range else header_col
        max_col = header_range.max_col if header_range else header_col

        for row in range(header_row + 1, min(reader.sheet.max_row, header_row + 3) + 1):
            for col in range(min_col, max_col + 1):
                value = reader.cell_value(row, col)
                percent = self._parse_percent(value)
                if percent != "":
                    return percent, as_cell(row, col)
        return "", ""

    @staticmethod
    def _parse_time_room(text: str) -> tuple[str, str, str]:
        if not text:
            return "", "", ""

        days: list[str] = []
        periods: list[str] = []
        rooms: list[str] = []

        pattern = re.compile(r"([월화수목금토일])\s*([0-9,\s]+?)\s*\[([^\]]+)\]")
        for match in pattern.finditer(text):
            day = match.group(1)
            period_numbers = re.findall(r"\d+", match.group(2))
            room = clean_text(match.group(3))
            days.append(day)
            periods.append(",".join(period_numbers))
            rooms.append(room)

        if days:
            return "|".join(days), "|".join(periods), "|".join(rooms)

        return "", clean_text(text), ""


class LectureExporter:
    def __init__(self, parser: LectureParser | None = None, logger: logging.Logger | None = None):
        self.logger = logger or logging.getLogger(self.__class__.__name__)
        self.parser = parser or LectureParser(logger=self.logger)

    def export(self, paths: Iterable[str | Path], output_path: str | Path = "lecture_data.csv") -> list[dict[str, Any]]:
        rows = []
        for path in paths:
            self.logger.info("parsing %s", path)
            rows.append(self.parser.parse_file(path))

        output_path = Path(output_path)
        with output_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            writer.writeheader()
            writer.writerows(rows)

        self.logger.info("saved %s (%d rows)", output_path, len(rows))
        return rows


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Parse Chungbuk University lecture plan XLSX files to CSV.")
    parser.add_argument(
        "xlsx",
        nargs="*",
        help="Input XLSX files. If omitted, all .xlsx files in the current folder are parsed.",
    )
    parser.add_argument("-o", "--output", default="extract_lecture.csv", help="Output CSV path")
    parser.add_argument("--debug", action="store_true", help="Print detailed cell-read logs")
    parser.add_argument("--dump-structure", action="store_true", help="Print sheet dimensions and merged ranges")
    parser.add_argument("--dump-rows", action="store_true", help="Print all non-empty cells before parsing")
    parser.add_argument("--dump-limit", type=int, default=70, help="Maximum row count for --dump-rows")
    parser.add_argument("--fill-merged", action="store_true", help="When dumping rows, fill merged cells with top-left values")
    return parser


def configure_logging(debug: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if debug else logging.INFO,
        format="%(levelname)s:%(name)s:%(message)s",
    )


def main() -> None:
    args = build_arg_parser().parse_args()
    configure_logging(args.debug)
    logger = logging.getLogger("lecture_excel_parser")
    xlsx_paths = args.xlsx or sorted(str(path) for path in Path.cwd().glob("*.xlsx") if not path.name.startswith("~$"))

    if not xlsx_paths:
        raise SystemExit(
            "No XLSX files found.\n"
            "Put the lecture .xlsx files in this folder and run again, or pass files explicitly:\n"
            "  python extract_lectures.py report_1.xlsx report_2.xlsx"
        )

    logger.info("input XLSX files: %d", len(xlsx_paths))

    if args.dump_structure or args.dump_rows:
        for xlsx_path in xlsx_paths:
            reader = LectureExcelReader(xlsx_path, logger=logger)
            print("\n" + "=" * 100)
            if args.dump_structure:
                print(reader.dump_structure())
            if args.dump_rows:
                print(reader.dump_all_rows(limit=args.dump_limit, fill_merged=args.fill_merged))

    exporter = LectureExporter(logger=logger)
    exporter.export(xlsx_paths, args.output)


if __name__ == "__main__":
    main()
